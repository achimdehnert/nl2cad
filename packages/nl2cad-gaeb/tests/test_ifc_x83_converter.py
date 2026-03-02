"""Tests fuer nl2cad.gaeb.converter.IFCX83Converter."""

from __future__ import annotations

import xml.etree.ElementTree as ET
from io import BytesIO

from nl2cad.gaeb.converter import IFCX83Converter


class TestIFCX83ConverterXML:
    def test_convert_returns_bytesio(self, sample_ifc_data):
        converter = IFCX83Converter()
        result = converter.convert_to_x83(sample_ifc_data, projekt_name="Test")
        assert isinstance(result, BytesIO)

    def test_convert_returns_valid_xml(self, sample_ifc_data):
        converter = IFCX83Converter()
        result = converter.convert_to_x83(sample_ifc_data, projekt_name="Test")
        root = ET.fromstring(result.read())
        assert root is not None

    def test_convert_contains_projekt_name(self, sample_ifc_data):
        converter = IFCX83Converter()
        result = converter.convert_to_x83(
            sample_ifc_data, projekt_name="Mein Neubau"
        )
        content = result.read().decode("utf-8")
        assert "Mein Neubau" in content

    def test_convert_has_positions_for_rooms(self, sample_ifc_data):
        converter = IFCX83Converter()
        result = converter.convert_to_x83(sample_ifc_data, projekt_name="Test")
        content = result.read().decode("utf-8")
        assert "Buero 1" in content
        assert "Buero 2" in content

    def test_convert_empty_ifc_data(self):
        converter = IFCX83Converter()
        result = converter.convert_to_x83({}, projekt_name="Leer")
        assert result is not None
        root = ET.fromstring(result.read())
        assert root is not None

    def test_convert_no_rooms_means_no_positions(self):
        converter = IFCX83Converter()
        ifc_data = {"project_name": "Test", "rooms": []}
        result = converter.convert_to_x83(ifc_data, projekt_name="Test")
        content = result.read().decode("utf-8")
        assert "Bodenbelag" not in content

    def test_convert_with_projekt_nummer(self, sample_ifc_data):
        converter = IFCX83Converter()
        result = converter.convert_to_x83(
            sample_ifc_data,
            projekt_name="Test",
            projekt_nummer="2026-001",
        )
        assert result is not None


class TestIFCX83ConverterExcel:
    def test_convert_to_excel_returns_readable_workbook(self, sample_ifc_data):
        from openpyxl import load_workbook

        converter = IFCX83Converter()
        result = converter.convert_to_excel(
            sample_ifc_data, projekt_name="Test"
        )
        wb = load_workbook(result)
        assert "Leistungsverzeichnis" in wb.sheetnames

    def test_convert_to_excel_contains_rooms(self, sample_ifc_data):
        from openpyxl import load_workbook

        converter = IFCX83Converter()
        result = converter.convert_to_excel(
            sample_ifc_data, projekt_name="Test"
        )
        wb = load_workbook(result)
        ws = wb["Leistungsverzeichnis"]
        values = [ws.cell(row, 2).value for row in range(1, ws.max_row + 1)]
        assert any("Buero 1" in str(v) for v in values if v)
