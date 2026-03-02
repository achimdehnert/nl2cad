"""Tests fuer nl2cad.gaeb.generator.GAEBGenerator."""

from __future__ import annotations

import xml.etree.ElementTree as ET
from decimal import Decimal

from nl2cad.gaeb.generator import GAEBGenerator
from nl2cad.gaeb.models import Leistungsverzeichnis, LosGruppe, Position


def _make_lv(projekt_name: str = "Testprojekt") -> Leistungsverzeichnis:
    return Leistungsverzeichnis(
        projekt_name=projekt_name,
        lose=[
            LosGruppe(
                oz="01",
                bezeichnung="Bodenbelaege",
                positionen=[
                    Position(
                        oz="01.001",
                        kurztext="Parkett",
                        menge=Decimal("50.0"),
                        einheit="m²",
                        einheitspreis=Decimal("60.00"),
                    ),
                    Position(
                        oz="01.002",
                        kurztext="Fliesen",
                        menge=Decimal("12.0"),
                        einheit="m²",
                        einheitspreis=Decimal("40.00"),
                    ),
                ],
            )
        ],
    )


class TestGAEBGeneratorXML:
    def test_generate_xml_returns_bytesio(self):
        from io import BytesIO

        gen = GAEBGenerator()
        output = gen.generate_xml(_make_lv())
        assert isinstance(output, BytesIO)

    def test_generated_xml_is_valid(self):
        gen = GAEBGenerator()
        output = gen.generate_xml(_make_lv())
        root = ET.fromstring(output.read())
        assert root is not None

    def test_generated_xml_contains_project_name(self):
        gen = GAEBGenerator()
        output = gen.generate_xml(_make_lv("Mein Projekt"))
        content = output.read().decode("utf-8")
        assert "Mein Projekt" in content

    def test_generated_xml_contains_positions(self):
        gen = GAEBGenerator()
        output = gen.generate_xml(_make_lv())
        content = output.read().decode("utf-8")
        assert "Parkett" in content
        assert "Fliesen" in content

    def test_generated_xml_contains_los(self):
        gen = GAEBGenerator()
        output = gen.generate_xml(_make_lv())
        content = output.read().decode("utf-8")
        assert "Bodenbelaege" in content

    def test_generated_xml_has_xml_declaration(self):
        gen = GAEBGenerator()
        output = gen.generate_xml(_make_lv())
        content = output.read().decode("utf-8")
        assert "<?xml" in content

    def test_empty_lv_generates_valid_xml(self):
        gen = GAEBGenerator()
        lv = Leistungsverzeichnis(projekt_name="Leer")
        output = gen.generate_xml(lv)
        root = ET.fromstring(output.read())
        assert root is not None

    def test_generate_xml_with_fixture(self, simple_lv):
        gen = GAEBGenerator()
        output = gen.generate_xml(simple_lv)
        root = ET.fromstring(output.read())
        assert root is not None


class TestGAEBGeneratorExcel:
    def test_generate_excel_returns_readable_workbook(self):
        from openpyxl import load_workbook

        gen = GAEBGenerator()
        output = gen.generate_excel(_make_lv())
        wb = load_workbook(output)
        assert "Leistungsverzeichnis" in wb.sheetnames

    def test_generate_excel_contains_header(self):
        from openpyxl import load_workbook

        gen = GAEBGenerator()
        output = gen.generate_excel(_make_lv())
        wb = load_workbook(output)
        ws = wb["Leistungsverzeichnis"]
        headers = [ws.cell(1, i).value for i in range(1, 7)]
        assert "OZ" in headers
        assert "Kurztext" in headers

    def test_generate_excel_contains_positions(self):
        from openpyxl import load_workbook

        gen = GAEBGenerator()
        output = gen.generate_excel(_make_lv())
        wb = load_workbook(output)
        ws = wb["Leistungsverzeichnis"]
        values = [ws.cell(row, 2).value for row in range(1, ws.max_row + 1)]
        assert "Parkett" in values

    def test_generate_excel_with_fixture(self, simple_lv):
        from openpyxl import load_workbook

        gen = GAEBGenerator()
        output = gen.generate_excel(simple_lv)
        wb = load_workbook(output)
        assert "Leistungsverzeichnis" in wb.sheetnames
